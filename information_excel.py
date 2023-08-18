import pandas as pd
from openpyxl import Workbook, load_workbook

archivos_a_verificar = 'list.csv'
hashtag_xlsx = 'hashtags.xlsx'
media_value_xlsx = 'media_value.xlsx'

df_hashtags = pd.read_excel(hashtag_xlsx, index_col=0)

resultados_por_marca = {}

# Leer la lista de archivos a verificar desde el archivo CSV
lista_archivos = pd.read_csv(archivos_a_verificar, header=None)[0]

# Leer los límites y recompensas desde el archivo 'media_value.xlsx'
limits_df = pd.read_excel(media_value_xlsx)

limits_df['From'] = pd.to_numeric(limits_df['From'], errors='coerce')
limits_df['To'] = pd.to_numeric(limits_df['To'], errors='coerce')

for archivo_csv in lista_archivos:
    # Leer el archivo CSV en un DataFrame de pandas
    df = pd.read_csv(archivo_csv)
    columna_texto = df.columns[0]
    columna_2 = df.columns[1]
    columna_3 = df.columns[2]
    columna_4 = df.columns[3]
    print(columna_4)

    for marca, hashtags in df_hashtags.iterrows():
        resultados_encontrados = []

        for idx, resultado in enumerate(df[columna_texto]):
            for hashtag in hashtags.dropna():
                if hashtag in resultado:
                    value_to_check_str = df.loc[idx, columna_4]
                    value_to_check = int(value_to_check_str.replace("'", "").replace(",", ""))
                    reward = None

                    # Verificar si el valor se encuentra dentro de los límites y obtener la recompensa
                    for _, row in limits_df.iterrows():
                        if row['From'] <= value_to_check <= row['To']:
                            reward = row['Amount']
                            break

                    resultados_encontrados.append((resultado, df.loc[idx, columna_2], df.loc[idx, columna_3], df.loc[idx, columna_4], reward))
        resultados_por_marca[(archivo_csv, marca)] = resultados_encontrados

# Crear un archivo Excel de salida y hojas de cálculo con nombres de archivo y marca
output_filename = 'Results.xlsx'
workbook = Workbook()
for (archivo_csv, marca), resultados in resultados_por_marca.items():
    worksheet = workbook.create_sheet(title=f'{marca}')
    worksheet.append(["Text", "Time", "Link", "Reach", "Reward"])  # Agregar la primera fila con encabezados
    for idx, (resultado, columna2, columna3, columna4, reward) in enumerate(resultados, start=2):  # Comenzar desde la fila 2 para los datos
        worksheet.cell(row=idx, column=1, value=resultado)
        worksheet.cell(row=idx, column=2, value=columna2)
        worksheet.cell(row=idx, column=3, value=columna3)
        worksheet.cell(row=idx, column=4, value=columna4)
        worksheet.cell(row=idx, column=5, value=reward)

# Guardar el archivo Excel
workbook.remove(workbook['Sheet'])  # Eliminar la hoja en blanco predeterminada
workbook.save(output_filename)
